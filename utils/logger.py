from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from config import Config

class VerificationLogger:
    @staticmethod
    def log_verification(img1_name, img2_name, verification_result):
        if not verification_result['is_match']:
            return
            
        try:
            # Create logs directory if it doesn't exist
            log_dir = os.path.dirname(Config.FILE_NAME)
            if log_dir and not os.path.exists(log_dir):
                os.makedirs(log_dir)
            
            try:
                workbook = load_workbook(Config.FILE_NAME)
                sheet = workbook.active
            except FileNotFoundError:
                print("Creating new Excel file...")
                workbook = Workbook()
                sheet = workbook.active
                # Add headers
                sheet.append(["Date & Time", "Image 1", "Image 2", "Cosine Similarity", 
                            "Euclidean Distance", "Verification Status"])
        
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Extract just the filename from the full paths
            img1_name = os.path.basename(img1_name)
            img2_name = os.path.basename(img2_name)
            
            sheet.append([
                current_time,
                img1_name,
                img2_name,
                round(float(verification_result['cosine_similarity']), 4),
                round(float(verification_result['euclidean_distance']), 4),
                verification_result['verification_status']
            ])
            
            try:
                workbook.save(Config.FILE_NAME)
                print(f"Results saved to {Config.FILE_NAME}")
            except PermissionError:
                print(f"Permission denied: Could not save to {Config.FILE_NAME}. Please close the Excel file if it's open.")
            except Exception as e:
                print(f"Error saving the Excel file: {e}")
                
        except Exception as e:
            print(f"Error in logging verification: {e}")